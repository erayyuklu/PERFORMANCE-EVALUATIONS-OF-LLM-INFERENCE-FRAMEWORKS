"""
Real agent tools for GAIA-compatible evaluation.
=================================================
These tools perform actual web searches, page fetching, sandboxed code
execution (via E2B), and local file reading — replacing the mock-tools
service when TOOL_MODE=real.
"""

import csv
import io
import logging
import os
import re

import httpx
from langchain_core.tools import tool

logger = logging.getLogger(__name__)

MAX_PAGE_CHARS = 20_000  # Truncation limit for visit_webpage output


@tool
async def web_search(query: str) -> str:
    """Search the web for a query. Returns top results with titles, URLs, and snippets."""
    from ddgs import DDGS

    try:
        with DDGS() as ddgs:
            results = list(ddgs.text(query, region="wt-wt", max_results=8))
        if not results:
            return "No search results found."
        parts = []
        for r in results:
            parts.append(f"**{r['title']}**\n{r['href']}\n{r['body']}")
        return "\n\n".join(parts)
    except Exception as exc:
        return f"Search error: {exc}"


@tool
async def visit_webpage(url: str) -> str:
    """Fetch a webpage and return its text content as readable markdown."""
    from bs4 import BeautifulSoup
    from markdownify import markdownify

    try:
        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
            resp = await client.get(
                url, headers={"User-Agent": "Mozilla/5.0 (GAIA-Agent)"}
            )
            resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        # Remove non-content elements
        for tag in soup(["script", "style", "nav", "footer", "header"]):
            tag.decompose()
        md = markdownify(str(soup), heading_style="ATX", strip=["img"])
        # Collapse excessive whitespace
        md = re.sub(r"\n{3,}", "\n\n", md).strip()
        if len(md) > MAX_PAGE_CHARS:
            md = md[:MAX_PAGE_CHARS] + "\n\n[...truncated]"
        return md if md else "Page returned no readable content."
    except Exception as exc:
        return f"Failed to fetch {url}: {exc}"


@tool
async def python_execute(code: str) -> str:
    """Execute Python code in a secure cloud sandbox. Returns stdout and stderr."""
    from e2b_code_interpreter import Sandbox

    try:
        sandbox = Sandbox()
        execution = sandbox.run_code(code)
        sandbox.kill()

        parts = []
        if execution.text:
            parts.append(execution.text)
        if execution.logs.stdout:
            parts.append("".join(execution.logs.stdout))
        if execution.logs.stderr:
            parts.append("[stderr]\n" + "".join(execution.logs.stderr))
        if execution.error:
            parts.append(f"[error] {execution.error.name}: {execution.error.value}")

        output = "\n".join(parts).strip()
        return output if output else "(no output)"
    except Exception as exc:
        return f"Execution error: {exc}"


@tool
async def read_document(file_path: str) -> str:
    """Read a local file and return its text content. Supports PDF, Excel, CSV, and plain text."""
    if not os.path.isfile(file_path):
        return f"File not found: {file_path}"

    ext = os.path.splitext(file_path)[1].lower()

    try:
        if ext == ".pdf":
            import pymupdf

            doc = pymupdf.open(file_path)
            pages = [page.get_text() for page in doc]
            doc.close()
            return "\n\n".join(pages).strip() or "(empty PDF)"

        if ext in (".xlsx", ".xls"):
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append("\t".join(str(c) if c is not None else "" for c in row))
                parts.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))
            wb.close()
            return "\n\n".join(parts).strip() or "(empty spreadsheet)"

        if ext == ".csv":
            with open(file_path, newline="", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                rows = ["\t".join(row) for row in reader]
            return "\n".join(rows).strip() or "(empty CSV)"

        # Default: plain text (txt, py, json, md, etc.)
        with open(file_path, encoding="utf-8", errors="replace") as f:
            content = f.read()
        return content.strip() or "(empty file)"

    except Exception as exc:
        return f"Error reading {file_path}: {exc}"


def get_real_tools():
    """Return all real agent tools."""
    return [web_search, visit_webpage, python_execute, read_document]